from flask import Flask, request, jsonify, send_file, send_from_directory, redirect, session
from flask_cors import CORS
from config import Config
from database import db
from models import Aluno, Aula, Contrato, ListaEspera, Relatorio, GoogleCalendarToken, StatusAluno, StatusAula
from datetime import datetime, date
import os
import json
from werkzeug.utils import secure_filename
from io import BytesIO
# Importações opcionais do Google Calendar
try:
    from google_calendar import get_google_credentials, create_google_calendar_event, update_google_calendar_event, delete_google_calendar_event
    from google_auth_oauthlib.flow import Flow
    from google.auth.transport.requests import Request
    GOOGLE_CALENDAR_AVAILABLE = True
except ImportError:
    GOOGLE_CALENDAR_AVAILABLE = False
    print("Aviso: Google Calendar não disponível. Instale as dependências com: pip install -r requirements.txt")

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = app.config['SECRET_KEY']
CORS(app)
db.init_app(app)

# Criar pastas necessárias
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/uploads', exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# ========== ROTA DE IMPORTAÇÃO ==========
@app.route('/api/importar-alunos', methods=['POST'])
def importar_alunos_endpoint():
    """Endpoint para importar alunos de uma planilha Excel"""
    try:
        from openpyxl import load_workbook
        
        # Caminho da planilha
        planilha_path = os.path.join('imports', 'Planilha_Aulas_Teacher_Nathy.xlsx')
        
        if not os.path.exists(planilha_path):
            return jsonify({'error': 'Planilha não encontrada'}), 404
        
        # Carregar planilha
        wb = load_workbook(planilha_path, data_only=True)
        ws = wb.active
        
        alunos_importados = 0
        alunos_atualizados = 0
        alunos_erro = 0
        erros = []
        
        # Processar linhas (começando da linha 2, linha 1 é cabeçalho)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            
            try:
                # Coluna A: Nome do aluno
                nome = str(row[0]).strip() if row[0] else None
                if not nome or nome.lower() in ['nome', 'aluno', '']:
                    continue
                
                # Coluna B: Valor Hora-Aula (R$)
                valor_hora = None
                if len(row) > 1 and row[1]:
                    try:
                        valor_hora = float(row[1])
                    except (ValueError, TypeError):
                        pass
                
                # Coluna C: Nº de Aulas no Mês
                frequencia_mensal = None
                if len(row) > 2 and row[2]:
                    try:
                        frequencia_mensal = int(float(row[2]))
                    except (ValueError, TypeError):
                        pass
                
                # Coluna E: Observações
                observacoes_planilha = None
                if len(row) > 4 and row[4]:
                    observacoes_planilha = str(row[4]).strip()
                
                # Montar observações
                observacoes_parts = []
                if valor_hora:
                    observacoes_parts.append(f"Valor hora/aula: R$ {valor_hora:.2f}")
                if frequencia_mensal:
                    observacoes_parts.append(f"Frequência mensal: {frequencia_mensal} aulas")
                if observacoes_planilha:
                    observacoes_parts.append(observacoes_planilha)
                
                observacoes = "Importado da planilha. " + ". ".join(observacoes_parts) if observacoes_parts else "Importado da planilha"
                
                # Verificar se aluno já existe
                aluno_existente = Aluno.query.filter_by(nome=nome).first()
                if aluno_existente:
                    # Atualizar aluno existente
                    aluno_existente.valor_hora_aula = valor_hora
                    aluno_existente.frequencia_mensal = frequencia_mensal
                    if observacoes_parts:
                        obs_atual = aluno_existente.observacoes or ""
                        aluno_existente.observacoes = f"{obs_atual}. Atualizado da planilha: {'. '.join(observacoes_parts)}" if obs_atual else observacoes
                    alunos_atualizados += 1
                    continue
                
                # Criar novo aluno
                aluno = Aluno(
                    nome=nome,
                    status=StatusAluno.ATIVO,
                    valor_hora_aula=valor_hora,
                    frequencia_mensal=frequencia_mensal,
                    observacoes=observacoes
                )
                
                db.session.add(aluno)
                alunos_importados += 1
                
            except Exception as e:
                alunos_erro += 1
                erros.append(f"Linha {row_num}: {str(e)}")
                continue
        
        # Commit das alterações
        db.session.commit()
        
        return jsonify({
            'message': 'Importação concluída',
            'alunos_importados': alunos_importados,
            'alunos_atualizados': alunos_atualizados,
            'erros': alunos_erro,
            'detalhes_erros': erros[:10]  # Limitar a 10 erros
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ========== ROTAS DE ALUNOS ==========
@app.route('/api/alunos', methods=['GET'])
def listar_alunos():
    try:
        alunos = Aluno.query.all()
        return jsonify([aluno.to_dict() for aluno in alunos]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alunos/<int:id>', methods=['GET'])
def obter_aluno(id):
    try:
        aluno = Aluno.query.get_or_404(id)
        return jsonify(aluno.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alunos', methods=['POST'])
def criar_aluno():
    try:
        data = request.get_json()
        aluno = Aluno(
            nome=data.get('nome'),
            email=data.get('email'),
            telefone=data.get('telefone'),
            endereco=data.get('endereco'),
            data_nascimento=datetime.fromisoformat(data['data_nascimento']).date() if data.get('data_nascimento') else None,
            observacoes=data.get('observacoes'),
            status=StatusAluno[data.get('status', 'ATIVO').upper()]
        )
        db.session.add(aluno)
        db.session.commit()
        return jsonify(aluno.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/alunos/<int:id>', methods=['PUT'])
def atualizar_aluno(id):
    try:
        aluno = Aluno.query.get_or_404(id)
        data = request.get_json()
        
        aluno.nome = data.get('nome', aluno.nome)
        aluno.email = data.get('email', aluno.email)
        aluno.telefone = data.get('telefone', aluno.telefone)
        aluno.endereco = data.get('endereco', aluno.endereco)
        if data.get('data_nascimento'):
            aluno.data_nascimento = datetime.fromisoformat(data['data_nascimento']).date()
        aluno.observacoes = data.get('observacoes', aluno.observacoes)
        if data.get('status'):
            aluno.status = StatusAluno[data['status'].upper()]
        
        db.session.commit()
        return jsonify(aluno.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/alunos/<int:id>', methods=['DELETE'])
def deletar_aluno(id):
    try:
        aluno = Aluno.query.get_or_404(id)
        db.session.delete(aluno)
        db.session.commit()
        return jsonify({'message': 'Aluno deletado com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ========== ROTAS DE AULAS ==========
@app.route('/api/aulas', methods=['GET'])
def listar_aulas():
    try:
        aluno_id = request.args.get('aluno_id', type=int)
        query = Aula.query
        if aluno_id:
            query = query.filter_by(aluno_id=aluno_id)
        aulas = query.order_by(Aula.data_aula.desc()).all()
        return jsonify([aula.to_dict() for aula in aulas]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/aulas/<int:id>', methods=['GET'])
def obter_aula(id):
    try:
        aula = Aula.query.get_or_404(id)
        return jsonify(aula.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/aulas', methods=['POST'])
def criar_aula():
    try:
        from datetime import timedelta
        data = request.get_json()
        
        # Verificar se deve repetir semanalmente
        repetir_semanal = data.get('repetir_semanal', False)
        num_semanas = data.get('num_semanas', 1) if repetir_semanal else 1
        frequencia_semanal = data.get('frequencia_semanal', 1) if repetir_semanal else 1
        dias_semana = data.get('dias_semana', [])  # Lista de dias da semana (0=segunda, 6=domingo)
        
        aulas_criadas = []
        data_inicio = datetime.fromisoformat(data['data_aula'])
        
        # Gerar grupo_aula_id se for repetir (para vincular aulas relacionadas)
        import hashlib
        grupo_aula_id = None
        if repetir_semanal:
            # Criar hash único baseado em aluno_id + dias da semana + horário
            hora_inicio = data_inicio.time()
            dias_semana_str = ','.join(map(str, sorted(dias_semana))) if dias_semana else str(data_inicio.weekday())
            grupo_str = f"{data.get('aluno_id')}_{dias_semana_str}_{hora_inicio}"
            grupo_aula_id = hashlib.md5(grupo_str.encode()).hexdigest()[:16]
        
        # Se não repetir, criar apenas uma aula
        if not repetir_semanal:
            aula = Aula(
                aluno_id=data.get('aluno_id'),
                data_aula=data_inicio,
                duracao=data.get('duracao', 60),
                conteudo=data.get('conteudo'),
                observacoes=data.get('observacoes'),
                status=StatusAula[data.get('status', 'AGENDADA').upper()],
                grupo_aula_id=grupo_aula_id
            )
            db.session.add(aula)
            db.session.flush()
            
            # Sincronizar com Google Calendar
            if GOOGLE_CALENDAR_AVAILABLE:
                token_record = GoogleCalendarToken.query.first()
                if token_record and data.get('sincronizar_google', True):
                    creds = get_google_credentials(app.config, token_record)
                    if creds:
                        aluno = Aluno.query.get(aula.aluno_id)
                        if aluno:
                            event_id = create_google_calendar_event(creds, aula, aluno.nome)
                            if event_id:
                                aula.google_event_id = event_id
            
            aulas_criadas.append(aula.to_dict())
        else:
            # Se repetir semanalmente
            if frequencia_semanal == 1:
                # 1x por semana: usar o dia da data inicial
                dia_semana_inicio = data_inicio.weekday()  # 0=segunda, 6=domingo
                
                for semana in range(num_semanas):
                    data_aula = data_inicio + timedelta(weeks=semana)
                    
                    aula = Aula(
                        aluno_id=data.get('aluno_id'),
                        data_aula=data_aula,
                        duracao=data.get('duracao', 60),
                        conteudo=data.get('conteudo'),
                        observacoes=data.get('observacoes'),
                        status=StatusAula[data.get('status', 'AGENDADA').upper()]
                    )
                    db.session.add(aula)
                    db.session.flush()
                    
                    # Sincronizar com Google Calendar
                    if GOOGLE_CALENDAR_AVAILABLE:
                        token_record = GoogleCalendarToken.query.first()
                        if token_record and data.get('sincronizar_google', True):
                            creds = get_google_credentials(app.config, token_record)
                            if creds:
                                aluno = Aluno.query.get(aula.aluno_id)
                                if aluno:
                                    event_id = create_google_calendar_event(creds, aula, aluno.nome)
                                    if event_id:
                                        aula.google_event_id = event_id
                    
                    aulas_criadas.append(aula.to_dict())
            else:
                # Múltiplas vezes por semana: usar os dias selecionados
                if not dias_semana or len(dias_semana) != frequencia_semanal:
                    return jsonify({'error': f'Selecione exatamente {frequencia_semanal} dia(s) da semana'}), 400
                
                # Obter hora da data inicial
                hora_inicio = data_inicio.time()
                dia_semana_inicio = data_inicio.weekday()  # 0=segunda, 6=domingo
                
                for semana in range(num_semanas):
                    for dia_semana in sorted(dias_semana):  # Ordenar dias para manter ordem
                        if semana == 0:
                            # Primeira semana: calcular diferença do dia inicial
                            dias_diferenca = (dia_semana - dia_semana_inicio) % 7
                            if dias_diferenca == 0:
                                # Mesmo dia da data inicial
                                data_aula = data_inicio
                            else:
                                # Outro dia da primeira semana
                                data_aula = data_inicio + timedelta(days=dias_diferenca)
                        else:
                            # Semanas seguintes: calcular a partir da data inicial
                            # Semanas completas + diferença do dia
                            dias_diferenca = (dia_semana - dia_semana_inicio) % 7
                            data_aula = data_inicio + timedelta(weeks=semana, days=dias_diferenca)
                        
                        # Garantir que a hora seja mantida
                        if isinstance(data_aula, datetime):
                            if data_aula.time() != hora_inicio:
                                data_aula = datetime.combine(data_aula.date(), hora_inicio)
                        
                        aula = Aula(
                            aluno_id=data.get('aluno_id'),
                            data_aula=data_aula,
                            duracao=data.get('duracao', 60),
                            conteudo=data.get('conteudo'),
                            observacoes=data.get('observacoes'),
                            status=StatusAula[data.get('status', 'AGENDADA').upper()],
                            grupo_aula_id=grupo_aula_id
                        )
                        db.session.add(aula)
                        db.session.flush()
                        
                        # Sincronizar com Google Calendar
                        if GOOGLE_CALENDAR_AVAILABLE:
                            token_record = GoogleCalendarToken.query.first()
                            if token_record and data.get('sincronizar_google', True):
                                creds = get_google_credentials(app.config, token_record)
                                if creds:
                                    aluno = Aluno.query.get(aula.aluno_id)
                                    if aluno:
                                        event_id = create_google_calendar_event(creds, aula, aluno.nome)
                                        if event_id:
                                            aula.google_event_id = event_id
                        
                        aulas_criadas.append(aula.to_dict())
        
        db.session.commit()
        
        # Retornar a última aula criada (ou todas se múltiplas)
        if len(aulas_criadas) == 1:
            return jsonify(aulas_criadas[0]), 201
        else:
            return jsonify({
                'message': f'{len(aulas_criadas)} aulas criadas com sucesso',
                'aulas': aulas_criadas,
                'total': len(aulas_criadas)
            }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/aulas/<int:id>', methods=['PUT'])
def atualizar_aula(id):
    try:
        aula = Aula.query.get_or_404(id)
        data = request.get_json()
        
        # Verificar se deve atualizar todas as aulas do grupo
        atualizar_grupo = data.get('atualizar_grupo', False)
        aulas_para_atualizar = [aula]
        
        if atualizar_grupo and aula.grupo_aula_id:
            # Buscar todas as aulas do mesmo grupo
            aulas_grupo = Aula.query.filter_by(
                grupo_aula_id=aula.grupo_aula_id,
                aluno_id=aula.aluno_id
            ).all()
            aulas_para_atualizar = aulas_grupo
        
        # Atualizar todas as aulas do grupo
        for aula_atualizar in aulas_para_atualizar:
            if data.get('aluno_id'):
                aula_atualizar.aluno_id = data['aluno_id']
            if data.get('data_aula'):
                aula_atualizar.data_aula = datetime.fromisoformat(data['data_aula'])
            if data.get('duracao'):
                aula_atualizar.duracao = data['duracao']
            if data.get('conteudo') is not None:
                aula_atualizar.conteudo = data['conteudo']
            if data.get('observacoes') is not None:
                aula_atualizar.observacoes = data['observacoes']
            if data.get('status'):
                aula_atualizar.status = StatusAula[data['status'].upper()]
            
            # Sincronizar com Google Calendar se estiver conectado
            if GOOGLE_CALENDAR_AVAILABLE:
                token_record = GoogleCalendarToken.query.first()
                if token_record and aula_atualizar.google_event_id and data.get('sincronizar_google', True):
                    creds = get_google_credentials(app.config, token_record)
                    if creds:
                        aluno = Aluno.query.get(aula_atualizar.aluno_id)
                        if aluno:
                            update_google_calendar_event(creds, aula_atualizar.google_event_id, aula_atualizar, aluno.nome)
        
        db.session.commit()
        
        if len(aulas_para_atualizar) > 1:
            return jsonify({
                'message': f'{len(aulas_para_atualizar)} aulas atualizadas com sucesso',
                'aula': aula.to_dict(),
                'total_atualizadas': len(aulas_para_atualizar)
            }), 200
        
        return jsonify(aula.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/aulas/<int:id>', methods=['DELETE'])
def deletar_aula(id):
    try:
        aula = Aula.query.get_or_404(id)
        
        # Deletar evento do Google Calendar se existir
        if GOOGLE_CALENDAR_AVAILABLE and aula.google_event_id:
            token_record = GoogleCalendarToken.query.first()
            if token_record:
                creds = get_google_credentials(app.config, token_record)
                if creds:
                    delete_google_calendar_event(creds, aula.google_event_id)
        
        db.session.delete(aula)
        db.session.commit()
        return jsonify({'message': 'Aula deletada com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ========== ROTAS DE CONTRATOS ==========
@app.route('/api/contratos', methods=['GET'])
def listar_contratos():
    try:
        aluno_id = request.args.get('aluno_id', type=int)
        query = Contrato.query
        if aluno_id:
            query = query.filter_by(aluno_id=aluno_id)
        contratos = query.order_by(Contrato.data_upload.desc()).all()
        return jsonify([contrato.to_dict() for contrato in contratos]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/contratos/<int:id>', methods=['GET'])
def obter_contrato(id):
    try:
        contrato = Contrato.query.get_or_404(id)
        return jsonify(contrato.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/contratos', methods=['POST'])
def upload_contrato():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['arquivo']
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo vazio'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            filename = timestamp + filename
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            contrato = Contrato(
                aluno_id=request.form.get('aluno_id', type=int),
                nome_arquivo=file.filename,
                caminho_arquivo=filepath,
                tipo_documento=request.form.get('tipo_documento', 'documento'),
                observacoes=request.form.get('observacoes')
            )
            db.session.add(contrato)
            db.session.commit()
            return jsonify(contrato.to_dict()), 201
        else:
            return jsonify({'error': 'Tipo de arquivo não permitido'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/contratos/<int:id>/download', methods=['GET'])
def download_contrato(id):
    try:
        contrato = Contrato.query.get_or_404(id)
        return send_file(contrato.caminho_arquivo, as_attachment=True, download_name=contrato.nome_arquivo)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/contratos/<int:id>', methods=['DELETE'])
def deletar_contrato(id):
    try:
        contrato = Contrato.query.get_or_404(id)
        filepath = contrato.caminho_arquivo
        db.session.delete(contrato)
        db.session.commit()
        
        # Deletar arquivo físico
        if os.path.exists(filepath):
            os.remove(filepath)
        
        return jsonify({'message': 'Contrato deletado com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ========== ROTAS DE LISTA DE ESPERA ==========
@app.route('/api/lista-espera', methods=['GET'])
def listar_lista_espera():
    try:
        pessoas = ListaEspera.query.order_by(ListaEspera.prioridade.desc(), ListaEspera.data_cadastro.asc()).all()
        return jsonify([pessoa.to_dict() for pessoa in pessoas]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/lista-espera/<int:id>', methods=['GET'])
def obter_lista_espera(id):
    try:
        pessoa = ListaEspera.query.get_or_404(id)
        return jsonify(pessoa.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/lista-espera', methods=['POST'])
def adicionar_lista_espera():
    try:
        data = request.get_json()
        pessoa = ListaEspera(
            nome=data.get('nome'),
            email=data.get('email'),
            telefone=data.get('telefone'),
            observacoes=data.get('observacoes'),
            prioridade=data.get('prioridade', 5),
            status=data.get('status', 'aguardando')
        )
        db.session.add(pessoa)
        db.session.commit()
        return jsonify(pessoa.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/lista-espera/<int:id>', methods=['PUT'])
def atualizar_lista_espera(id):
    try:
        pessoa = ListaEspera.query.get_or_404(id)
        data = request.get_json()
        
        pessoa.nome = data.get('nome', pessoa.nome)
        pessoa.email = data.get('email', pessoa.email)
        pessoa.telefone = data.get('telefone', pessoa.telefone)
        pessoa.observacoes = data.get('observacoes', pessoa.observacoes)
        pessoa.prioridade = data.get('prioridade', pessoa.prioridade)
        pessoa.status = data.get('status', pessoa.status)
        
        db.session.commit()
        return jsonify(pessoa.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/lista-espera/<int:id>/ativar', methods=['POST'])
def ativar_lista_espera(id):
    try:
        pessoa = ListaEspera.query.get_or_404(id)
        data = request.get_json()
        
        # Criar aluno a partir da lista de espera
        aluno = Aluno(
            nome=pessoa.nome,
            email=pessoa.email,
            telefone=pessoa.telefone,
            observacoes=f"Convertido da lista de espera. {pessoa.observacoes or ''}",
            status=StatusAluno.ATIVO
        )
        db.session.add(aluno)
        
        # Atualizar status da lista de espera
        pessoa.status = 'convertido'
        db.session.commit()
        
        return jsonify({
            'message': 'Pessoa convertida para aluno com sucesso',
            'aluno': aluno.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/lista-espera/<int:id>', methods=['DELETE'])
def deletar_lista_espera(id):
    try:
        pessoa = ListaEspera.query.get_or_404(id)
        db.session.delete(pessoa)
        db.session.commit()
        return jsonify({'message': 'Registro deletado com sucesso'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ========== ROTAS DE RELATÓRIOS ==========
@app.route('/api/relatorios', methods=['GET'])
def listar_relatorios():
    try:
        aluno_id = request.args.get('aluno_id', type=int)
        query = Relatorio.query
        if aluno_id:
            query = query.filter_by(aluno_id=aluno_id)
        relatorios = query.order_by(Relatorio.data_criacao.desc()).all()
        return jsonify([relatorio.to_dict() for relatorio in relatorios]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/relatorios/gerar', methods=['POST'])
def gerar_relatorio():
    try:
        data = request.get_json()
        aluno_id = data.get('aluno_id')
        periodo_inicio = datetime.fromisoformat(data['periodo_inicio']).date() if data.get('periodo_inicio') else None
        periodo_fim = datetime.fromisoformat(data['periodo_fim']).date() if data.get('periodo_fim') else None
        
        # Buscar dados do aluno e aulas
        aluno = Aluno.query.get_or_404(aluno_id)
        query_aulas = Aula.query.filter_by(aluno_id=aluno_id)
        
        if periodo_inicio:
            query_aulas = query_aulas.filter(Aula.data_aula >= datetime.combine(periodo_inicio, datetime.min.time()))
        if periodo_fim:
            query_aulas = query_aulas.filter(Aula.data_aula <= datetime.combine(periodo_fim, datetime.max.time()))
        
        aulas = query_aulas.all()
        
        # Calcular estatísticas
        total_aulas = len(aulas)
        aulas_realizadas = len([a for a in aulas if a.status == StatusAula.REALIZADA])
        aulas_agendadas = len([a for a in aulas if a.status == StatusAula.AGENDADA])
        
        # Criar relatório
        relatorio_data = {
            'aluno': aluno.to_dict(),
            'periodo': {
                'inicio': periodo_inicio.isoformat() if periodo_inicio else None,
                'fim': periodo_fim.isoformat() if periodo_fim else None
            },
            'estatisticas': {
                'total_aulas': total_aulas,
                'aulas_realizadas': aulas_realizadas,
                'aulas_agendadas': aulas_agendadas,
                'taxa_frequencia': (aulas_realizadas / total_aulas * 100) if total_aulas > 0 else 0
            },
            'aulas': [aula.to_dict() for aula in aulas]
        }
        
        relatorio = Relatorio(
            aluno_id=aluno_id,
            titulo=data.get('titulo', f'Relatório - {aluno.nome}'),
            periodo_inicio=periodo_inicio,
            periodo_fim=periodo_fim,
            conteudo=json.dumps(relatorio_data, ensure_ascii=False)
        )
        db.session.add(relatorio)
        db.session.commit()
        
        return jsonify(relatorio.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/relatorios/<int:id>', methods=['GET'])
def obter_relatorio(id):
    try:
        relatorio = Relatorio.query.get_or_404(id)
        relatorio_dict = relatorio.to_dict()
        if relatorio.conteudo:
            relatorio_dict['dados'] = json.loads(relatorio.conteudo)
        return jsonify(relatorio_dict), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rota para servir o index.html
@app.route('/')
def index():
    return send_from_directory('templates', 'index.html')

# ========== ROTA DE ANIVERSÁRIOS ==========
@app.route('/api/aniversarios/planilha', methods=['GET'])
def gerar_planilha_aniversarios():
    """Gera uma planilha Excel com os aniversários dos alunos"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from calendar import month_name
        
        # Buscar alunos com data de nascimento
        alunos = Aluno.query.filter(Aluno.data_nascimento.isnot(None)).all()
        
        if not alunos:
            return jsonify({'error': 'Nenhum aluno com data de nascimento cadastrada'}), 404
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Aniversários"
        
        # Cabeçalhos
        headers = ['Nome', 'Data de Nascimento', 'Idade', 'Mês', 'Dia', 'Próximo Aniversário', 'Dias até Aniversário', 'Email', 'Telefone']
        ws.append(headers)
        
        # Estilizar cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Processar alunos e calcular informações
        hoje = date.today()
        alunos_com_aniversario = []
        
        for aluno in alunos:
            if aluno.data_nascimento:
                data_nasc = aluno.data_nascimento
                idade = hoje.year - data_nasc.year - ((hoje.month, hoje.day) < (data_nasc.month, data_nasc.day))
                
                # Calcular próximo aniversário
                proximo_aniversario = date(hoje.year, data_nasc.month, data_nasc.day)
                if proximo_aniversario < hoje:
                    proximo_aniversario = date(hoje.year + 1, data_nasc.month, data_nasc.day)
                
                dias_ate_aniversario = (proximo_aniversario - hoje).days
                
                alunos_com_aniversario.append({
                    'aluno': aluno,
                    'idade': idade,
                    'proximo_aniversario': proximo_aniversario,
                    'dias_ate': dias_ate_aniversario,
                    'mes': data_nasc.month,
                    'dia': data_nasc.day
                })
        
        # Ordenar por mês e dia
        alunos_com_aniversario.sort(key=lambda x: (x['mes'], x['dia']))
        
        # Adicionar dados
        for item in alunos_com_aniversario:
            aluno = item['aluno']
            row = [
                aluno.nome,
                aluno.data_nascimento.strftime('%d/%m/%Y'),
                item['idade'],
                month_name[item['mes']],
                item['dia'],
                item['proximo_aniversario'].strftime('%d/%m/%Y'),
                item['dias_ate'],
                aluno.email or '',
                aluno.telefone or ''
            ]
            ws.append(row)
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 30,  # Nome
            'B': 18,  # Data de Nascimento
            'C': 10,  # Idade
            'D': 15,  # Mês
            'E': 10,  # Dia
            'F': 20,  # Próximo Aniversário
            'G': 20,  # Dias até Aniversário
            'H': 30,  # Email
            'I': 18   # Telefone
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo com data atual
        filename = f"Aniversarios_{hoje.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========== ROTAS DO GOOGLE CALENDAR ==========
@app.route('/api/google/authorize', methods=['GET'])
def google_authorize():
    """Inicia o fluxo de autenticação OAuth do Google"""
    try:
        if not GOOGLE_CALENDAR_AVAILABLE:
            return jsonify({'error': 'Google Calendar não disponível. Instale as dependências: pip install -r requirements.txt'}), 400
        if not app.config['GOOGLE_CLIENT_ID'] or not app.config['GOOGLE_CLIENT_SECRET']:
            return jsonify({'error': 'Google Calendar não configurado. Configure GOOGLE_CLIENT_ID e GOOGLE_CLIENT_SECRET'}), 400
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": app.config['GOOGLE_CLIENT_ID'],
                    "client_secret": app.config['GOOGLE_CLIENT_SECRET'],
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [app.config['GOOGLE_REDIRECT_URI']]
                }
            },
            scopes=app.config['GOOGLE_SCOPES'],
            redirect_uri=app.config['GOOGLE_REDIRECT_URI']
        )
        
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )
        
        session['state'] = state
        return jsonify({'authorization_url': authorization_url}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/google/oauth2callback', methods=['GET'])
def google_oauth2callback():
    """Callback do OAuth do Google"""
    try:
        if not GOOGLE_CALENDAR_AVAILABLE:
            return jsonify({'error': 'Google Calendar não disponível'}), 400
        state = session.get('state')
        code = request.args.get('code')
        
        if not code:
            return jsonify({'error': 'Código de autorização não recebido'}), 400
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": app.config['GOOGLE_CLIENT_ID'],
                    "client_secret": app.config['GOOGLE_CLIENT_SECRET'],
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [app.config['GOOGLE_REDIRECT_URI']]
                }
            },
            scopes=app.config['GOOGLE_SCOPES'],
            redirect_uri=app.config['GOOGLE_REDIRECT_URI'],
            state=state
        )
        
        flow.fetch_token(code=code)
        credentials = flow.credentials
        
        # Salvar ou atualizar token
        token_record = GoogleCalendarToken.query.first()
        if token_record:
            token_record.token = json.dumps({
                'token': credentials.token,
                'refresh_token': credentials.refresh_token,
                'token_uri': credentials.token_uri,
                'client_id': credentials.client_id,
                'client_secret': credentials.client_secret,
                'scopes': credentials.scopes
            })
            token_record.refresh_token = credentials.refresh_token
            token_record.token_uri = credentials.token_uri
            token_record.client_id = credentials.client_id
            token_record.client_secret = credentials.client_secret
            token_record.scopes = ','.join(credentials.scopes) if credentials.scopes else None
        else:
            token_record = GoogleCalendarToken(
                token=json.dumps({
                    'token': credentials.token,
                    'refresh_token': credentials.refresh_token,
                    'token_uri': credentials.token_uri,
                    'client_id': credentials.client_id,
                    'client_secret': credentials.client_secret,
                    'scopes': credentials.scopes
                }),
                refresh_token=credentials.refresh_token,
                token_uri=credentials.token_uri,
                client_id=credentials.client_id,
                client_secret=credentials.client_secret,
                scopes=','.join(credentials.scopes) if credentials.scopes else None
            )
            db.session.add(token_record)
        
        db.session.commit()
        
        # Redirecionar para a página de sucesso
        return redirect('http://localhost:8000/#/aulas?google_connected=true')
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/google/status', methods=['GET'])
def google_status():
    """Verifica se o Google Calendar está conectado"""
    try:
        if not GOOGLE_CALENDAR_AVAILABLE:
            return jsonify({'connected': False, 'message': 'Google Calendar não disponível'}), 200
        token_record = GoogleCalendarToken.query.first()
        if token_record:
            creds = get_google_credentials(app.config, token_record)
            if creds and creds.valid:
                return jsonify({'connected': True, 'message': 'Google Calendar conectado'}), 200
        return jsonify({'connected': False, 'message': 'Google Calendar não conectado'}), 200
    except Exception as e:
        return jsonify({'connected': False, 'error': str(e)}), 200

@app.route('/api/google/disconnect', methods=['POST'])
def google_disconnect():
    """Desconecta o Google Calendar"""
    try:
        if not GOOGLE_CALENDAR_AVAILABLE:
            return jsonify({'error': 'Google Calendar não disponível'}), 400
        token_record = GoogleCalendarToken.query.first()
        if token_record:
            db.session.delete(token_record)
            db.session.commit()
        return jsonify({'message': 'Google Calendar desconectado'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Rota de health check para verificar se backend está disponível
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'backend': True}), 200

# Rota para servir arquivos estáticos
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=8000)
