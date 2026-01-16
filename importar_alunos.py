"""
Script para importar alunos de uma planilha Excel para o sistema
"""
import sys
import os
from openpyxl import load_workbook
from datetime import datetime

# Adicionar o diretório atual ao path para importar os módulos
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app
from database import db
from models import Aluno, StatusAluno

def importar_alunos(planilha_path):
    """
    Importa alunos de uma planilha Excel
    
    Estrutura esperada da planilha:
    - Coluna A: Nome do aluno
    - Coluna B: Valor da hora aula (opcional)
    - Coluna C+: Frequência de aulas por mês (opcional)
    """
    print(f"Carregando planilha: {planilha_path}")
    
    # Carregar planilha
    try:
        wb = load_workbook(planilha_path, data_only=True)
        ws = wb.active
        print(f"Planilha carregada. Aba ativa: {ws.title}")
    except Exception as e:
        print(f"Erro ao carregar planilha: {e}")
        return
    
    # Mostrar estrutura da planilha
    print("\nPrimeiras linhas da planilha:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        print(f"Linha {i}: {row}")
    
    # Detectar cabeçalho (primeira linha)
    cabecalho = [cell.value for cell in ws[1]]
    print(f"\nCabeçalho detectado: {cabecalho}")
    
    alunos_importados = 0
    alunos_erro = 0
    
    with app.app_context():
        # Processar linhas (começando da linha 2, assumindo que linha 1 é cabeçalho)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Pular linhas vazias
            if not row or not row[0]:
                continue
            
            try:
                nome = str(row[0]).strip() if row[0] else None
                if not nome or nome.lower() in ['nome', 'aluno', '']:
                    continue
                
                # Extrair valor da hora aula (se disponível)
                valor_hora = None
                if len(row) > 1 and row[1]:
                    try:
                        valor_hora = float(row[1])
                    except (ValueError, TypeError):
                        pass
                
                # Verificar se aluno já existe
                aluno_existente = Aluno.query.filter_by(nome=nome).first()
                if aluno_existente:
                    print(f"Aluno '{nome}' já existe. Pulando...")
                    continue
                
                # Criar novo aluno
                aluno = Aluno(
                    nome=nome,
                    status=StatusAluno.ATIVO,
                    observacoes=f"Importado da planilha. Valor hora/aula: R$ {valor_hora:.2f}" if valor_hora else "Importado da planilha"
                )
                
                db.session.add(aluno)
                alunos_importados += 1
                print(f"✓ Aluno '{nome}' adicionado")
                
            except Exception as e:
                alunos_erro += 1
                print(f"✗ Erro ao processar linha {row_num}: {e}")
                continue
        
        # Commit das alterações
        try:
            db.session.commit()
            print(f"\n{'='*50}")
            print(f"Importação concluída!")
            print(f"Alunos importados: {alunos_importados}")
            print(f"Erros: {alunos_erro}")
            print(f"{'='*50}")
        except Exception as e:
            db.session.rollback()
            print(f"\nErro ao salvar no banco de dados: {e}")

if __name__ == "__main__":
    planilha_path = os.path.join("imports", "Planilha_Aulas_Teacher_Nathy.xlsx")
    
    if not os.path.exists(planilha_path):
        print(f"Erro: Planilha não encontrada em {planilha_path}")
        sys.exit(1)
    
    importar_alunos(planilha_path)

